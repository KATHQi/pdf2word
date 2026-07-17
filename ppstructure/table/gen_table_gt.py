import os
import json
import argparse
import openpyxl
from openpyxl.utils import range_boundaries
from html import escape

def parse_paddle_label(label_path):
    data_map = {}
    if not os.path.exists(label_path): return data_map
    with open(label_path, 'r', encoding='utf-8') as f:
        for line in f:
            parts = line.strip().split('\t', 1)
            if len(parts) < 2: continue
            img_path = parts[0]
            labels = json.loads(parts[1])
            data_map[img_path] = labels
    return data_map

def excel_to_structure(excel_path):
    """提取Excel的HTML结构 tokens，格式与PPOCRLabel exportJSON一致"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # 构建 html_list，初始全为 'td'
    html_list = [['td'] * max_col for _ in range(max_row)]

    # 处理合并单元格（openpyxl: 1-indexed, 转成 0-indexed）
    for m_range in ws.merged_cells.ranges:
        min_col, min_row, max_col_b, max_row_b = range_boundaries(str(m_range))
        sr, er, sc, ec = min_row - 1, max_row_b, min_col - 1, max_col_b
        for i in range(sr, er):
            for j in range(sc, ec):
                html_list[i][j] = None
        html_list[sr][sc] = ''
        if ec - sc > 1:
            html_list[sr][sc] += ' colspan={}'.format(ec - sc)
        if er - sr > 1:
            html_list[sr][sc] += ' rowspan={}'.format(er - sr)

    # 转换为 token 列表（与 PPOCRLabel convert_token 逻辑相同）
    token_list = ['<tbody>']
    for row in html_list:
        token_list.append('<tr>')
        for col in row:
            if col is None:
                continue
            elif col == 'td':
                token_list.extend(['<td>', '</td>'])
            else:
                token_list.append('<td')
                if 'colspan' in col:
                    _, n = col.split('colspan=')
                    token_list.append(' colspan="{}"'.format(n.strip().split()[0]))
                if 'rowspan' in col:
                    _, n = col.split('rowspan=')
                    token_list.append(' rowspan="{}"'.format(n.strip().split()[0]))
                token_list.extend(['>', '</td>'])
        token_list.append('</tr>')
    token_list.append('</tbody>')

    return token_list

def rebuild_html(html_tokens, cells):
    """与 functions.py rebuild_html_from_ppstructure_label 完全相同的逻辑"""
    html_code = html_tokens.copy()
    to_insert = [i for i, tag in enumerate(html_code) if tag in ('<td>', '>')]
    for i, cell in zip(to_insert[::-1], cells[::-1]):
        if cell['tokens']:
            text = ''.join(
                escape(t) if len(t) == 1 else t
                for t in cell['tokens']
            )
            html_code.insert(i + 1, text)
    return '<html><body><table>{}</table></body></html>'.format(''.join(html_code))


def main(args):
    bbox_data = parse_paddle_label(args.bbox_txt)

    gt_lines = []

    for img_path, cell_bboxes in bbox_data.items():
        # 获取Excel结构 (假设excel与图片同名)
        excel_name = os.path.basename(img_path).split('.')[0] + ".xlsx"
        excel_path = os.path.join(args.excel_dir, excel_name)
        if not os.path.exists(excel_path):
            continue

        html_tokens = excel_to_structure(excel_path)

        # 直接从 bbox_txt 的 transcription 字段取文字，按 (y, x) 排序与 Excel 结构对齐
        # 与 functions.py dataset_transform 相同做法，无需 IOU 匹配
        sorted_cells = sorted(cell_bboxes, key=lambda a: (a['points'][0][1], a['points'][0][0]))
        cells = []
        for anno in sorted_cells:
            cells.append({
                'tokens': list(anno.get('transcription', '')),
                'bbox': anno['points']
            })

        gt_html = rebuild_html(html_tokens, cells)

        res = {
            "filename": os.path.basename(img_path),
            "html": {
                "structure": {"tokens": html_tokens},
                "cells": cells
            },
            "gt": gt_html
        }
        gt_lines.append(json.dumps(res, ensure_ascii=False))

    # 确保输出路径指向文件而不是目录
    if os.path.isdir(args.output):
        args.output = os.path.join(args.output, "gt.txt")
        print(f"Warning: Output path is a directory, using {args.output} as default file name.")

    # 确保输出目录存在
    out_dir = os.path.dirname(args.output)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir)
        
    with open(args.output, 'w', encoding='utf-8') as f:
        f.write("\n".join(gt_lines))
    print(f"Save GT to {args.output}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--bbox_txt", required=True, help="表格单元格标注文件（含 transcription 字段）")
    parser.add_argument("--excel_dir", required=True, help="存放xlsx文件的目录")
    parser.add_argument("--output", default="gt.txt", help="输出的标注文件")
    main(parser.parse_args())